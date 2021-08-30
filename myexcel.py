import pyodbc, os
from openpyxl import load_workbook

server = 'servsrv\mssql2016' 
database = 'VesselsDB'
username = 'sa' 
password = os.environ['SERVSRVPASS'] 
cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)
cursor = cnxn.cursor()

wb = load_workbook('a.xlsx')
results = []
ws = wb.worksheets[0]
for row in ws.iter_rows():
  results.append([cell.value for cell in row])

#delete first item of first row
#del results[0][0]
myyears = results[0][1:]

#now delete first row
myrecords = results[1:]

for index, myyear in enumerate(myyears):
  for myrecord in myrecords:
    cursor.execute("SELECT id from vessels_vessel where name='" + myrecord[0] + "';")
    row = cursor.fetchone()
    #print("insert into vessels_marketvalue(vessel_id, mvdate, mvprice) values (?, ?, ?)", row[0], str(myyear)+'-12-31', myrecord[index+1])
    cursor.execute("insert into vessels_marketvalue(vessel_id, mvdate, mvprice) values (?, ?, ?)", row[0], str(myyear)+'-12-31', myrecord[index+1])
    cnxn.commit()
