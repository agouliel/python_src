# https://pmc.ncbi.nlm.nih.gov/articles/PMC8011774/
# https://icm-experimental.springeropen.com/articles/10.1186/s40635-021-00421-8

from openpyxl import load_workbook
import csv

def pmsa(cvp_rap, map_in, co, age, height, weight, cvpb=0, mapb=0, cob=0,
  sv=0, svr=0, svv=0, sap=0, dap=0, hr=0, ppv=0
):
  a = 0.96
  b = 0.04
  c_nom = 0.96 * 0.038 * (94.17 + 0.193*age) # 0.96 is missing from the above link
  c_denom = 4.5 * (0.99**(age-15)) * 0.007184 * (height**0.725) * (weight**0.425)
  c = c_nom / c_denom
  pmsa = a*cvp_rap + b*map_in + c*co
  pmsab = a*cvpb + b*mapb + c*cob

  pvr = pmsa - cvp_rap # this is also called VRdP
  eh = pvr/pmsa

  rvr = pvr/co
  power = co * (map_in - cvp_rap) * 0.0022
  cpo = (co * map_in) / 451

  if pmsa != pmsab:
    epower = ((co * (map_in - cvp_rap)) - (cob * (mapb - cvpb))) * 0.0022 / (pmsa - pmsab)
    evol = (pmsa - cvp_rap) - (pmsab - cvpb) / (pmsa - pmsab)
  else:
    epower = 0
    evol = 0
  
  cart = sv / (sap - dap) if (sap - dap) != 0 else 0
  rart = map_in / (sv * hr) if (sv * hr) != 0 else 0
  rven = svr * 0.038
  ea = map_in / sv if sv != 0 else 0
  eadyn = ppv / svv if svv != 0 else 0

  return (round(pmsa,2), round(pvr,2), round(eh,2), round(rvr,2), round(power,2), round(cpo,2),
    round(epower,2), round(evol,2),
    round(cart,2), round(rart,2), round(rven,2), round(ea,2), round(eadyn,2),
  )

data = []
data.append(['CVP', 'MAP', 'CO', 'AGE', 'HEIGHT', 'WEIGHT', 'PMSA', 'PVR', 'EH', 'RVR', 'POWER', 'CPO']) # header

try:
  wb = load_workbook('thanos.xlsx')
  ws = wb.worksheets[0]
  for xlrow in ws.iter_rows():
    row = [cell.value for cell in xlrow]
    if row[0] and row[0]!='CVP': # exclude empty rows and header rows
      cvp = float(row[0])
      map_file = float(row[1])
      co = float(row[2])
      age = float(row[3])
      height = float(row[4])
      weight = float(row[5])
      result = pmsa(cvp, map_file, co, age, height, weight)
      #print(cvp, '\t', map_file, '\t', co, '\t', age, '\t', height, '\t', weight, '\t', result[0], '\t', result[1], '\t', result[2])
      data.append([cvp, map_file, co, age, height, weight, result[0], result[1], result[2], result[3], result[4], result[5]])

  with open('thanos.csv', 'w', newline='') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerows(data)
except:
  pass