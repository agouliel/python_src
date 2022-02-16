import streamlit as st
import pyodbc, os, platform, sys, os.path
from openpyxl import load_workbook

myexecute = False

header_table = "MRV_Header"
records_table = "MRV_Records"

myserver = 'servsrv\mssql2016' 
mydb = 'Technical'

try:
    myusername = os.environ['SERVDBUSER']
    mypassword = os.environ['SERVDBPASS']
except:
    myusername = input('Username: ')
    mypassword = input('Password: ')

cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + myserver + ';DATABASE=' + mydb + ';UID=' + myusername + ';PWD=' + mypassword)

cursor = cnxn.cursor()

if platform.platform().startswith('mac'):
    folder_str = '/Volumes/Drawings/MRV DATA/MRV FORM 01'
else:
    folder_str = r'\\192.168.0.11\Drawings\MRV DATA\MRV FORM 01'

os.chdir(folder_str)
listdir_result = os.listdir()

###############################

def process_file(xlfile, mytimestamp, mrvid=''):
  # delete temp table
  stmt = "delete from mrvtemp;"
  cursor.execute(stmt)

  try:
    # read from the excel file
    wb = load_workbook(xlfile, data_only=True)

    ws = wb.worksheets[0]

    # perform a loop over the excel rows
    for xlrow in ws.iter_rows():
      row = [cell.value for cell in xlrow]

      # insert each row into temp table
      stmt = f"""insert into mrvtemp(f1, f2, f3, f4, f5, f6, f7, f8, f9, f10, f11, f12, f13, f14, f15, f16, f17, f18, f19, f20, f21, f22)
      values('{row[0]}','{row[1]}','{row[2]}','{row[3]}','{row[4]}','{row[5]}','{row[6]}','{row[7]}','{row[8]}','{row[9]}','{row[10]}','{row[11]}',
      '{row[12]}','{row[13]}','{row[14]}','{row[15]}','{row[16]}','{row[17]}','{row[18]}','{row[19]}','{row[20]}','{row[21]}');"""
      cursor.execute(stmt)

    # check if file is correct
    stmt = "SELECT f2 FROM MRVTemp where f1 = 'FORM MRV 01'"
    cursor.execute(stmt)
    dbrow = cursor.fetchone()
    vesseltest = dbrow[0]

    # insert header record
    if mrvid:
      stmt = f"SET IDENTITY_INSERT {header_table} ON"
      cursor.execute(stmt)

      stmt = f"""with
a as (select convert(date, f2, 102) "Date" FROM MRVTemp where f1 = 'Date:'),
b as (SELECT f2 "Vessel" FROM MRVTemp where f1 = 'FORM MRV 01'),
c as (select substring(f1,6,len(f1)+1-6) "rev" from MRVTemp where f1 like 'REV%'),
d as (select f10 "Remark" from MRVTemp where f9 = 'Remark:')
insert into {header_table}(mrv_id, report_date, vessel, remarks, origin_file_name, rev, origin_file_date)
select {mrvid}, a.Date, b.Vessel, d.Remark, '{xlfile}', c.rev, '{mytimestamp}'
from a,b,c,d;"""

      cursor.execute(stmt)

      stmt = f"SET IDENTITY_INSERT {header_table} OFF"
      cursor.execute(stmt)

    else:
      stmt = f"""with
a as (select convert(date, f2, 102) "Date" FROM MRVTemp where f1 = 'Date:'),
b as (SELECT f2 "Vessel" FROM MRVTemp where f1 = 'FORM MRV 01'),
c as (select substring(f1,6,len(f1)+1-6) "rev" from MRVTemp where f1 like 'REV%'),
d as (select f10 "Remark" from MRVTemp where f9 = 'Remark:')
insert into {header_table}(report_date, vessel, remarks, origin_file_name, rev, origin_file_date)
select a.Date, b.Vessel, d.Remark, '{xlfile}', c.rev, '{mytimestamp}'
from a,b,c,d;"""

      cursor.execute(stmt)

    # insert detail records (first get mrv_id from previously inserted header record)
    stmt = f"""with
a as (select mrv_id from {header_table} where origin_file_name = '{xlfile}'), 

b as (select t.Tank_ID, 
		m.f3 "BDN Number of fuel in tank", 
		m.f4 "FUEL TYPE (as per ISO 8217)", 
		cast(replace(m.f5,'None','0') as float) sulphur,
		cast(replace(m.f6,'None','0') as float) "SOUNDING", 
		cast(replace(m.f8,'None','0') as float) "TAPE READING",
		cast(replace(m.f10,'None','0') as float) "BOB READING", 
		cast(replace(m.f12,'None','0') as float) "ULLAGE", 
		cast(replace(m.f13,'None','0') as float) "TANK TEMP (°C)",
		cast(replace(m.f14,'None','0') as float) "TOTAL OBSERVED VOLUME (T.O.V. (m3))", 
		cast(replace(m.f16,'None','0') as float) "DENSITY AT 15oC (kg/lt)",
		cast(replace(m.f18,'None','0') as float) "Volume Correction Factor",
		cast(replace(m.f20,'None','0') as float) "GROSS STANDARD VOLUME (G.S.V. (m3))", 
		cast(replace(m.f21,'None','0') as float) "WEIGHT CORRECTION FACTOR (W.C.F.)",
		cast(replace(m.f22,'None','0') as float) "Weight"
	from MRVTemp m, MRV_Tanks t
	where m.f1 = t.Tank_Descr
)
insert into {records_table}
select *
from a,b;"""

    cursor.execute(stmt)

    cnxn.commit()

  except Exception as e:
    print(xlfile, e)
    stmt = f"""insert into MRV_Error_Log SELECT GETDATE(), '{xlfile}', '{str(e).replace("'","")}';"""
    cursor.execute(stmt)
    cnxn.commit()

######################################

def myfilter(infile):
  return infile.lower().endswith(('.xlsx','.xlsm')) and not infile.startswith('~')

######################################

def process_mrv():
    stmt = 'select origin_file_name, origin_file_date from mrv_header'
    cursor.execute(stmt)
    result = cursor.fetchall()
    processed_files_list = []
    for row in result:
        processed_files_list.append(row[0].lower()+'@@@'+row[1])

    #return listdir_result[0]+processed_files_list[0]

    # get a list of all current filenames and timestamps in the folder
    files_and_ts = []
    for xlfile in listdir_result:
        if myfilter(xlfile):
            try:
                files_and_ts.append(xlfile.lower()+'@@@'+str(os.path.getmtime(xlfile)).split('.')[0]+'.0')
            except:
                pass

    # if files have been removed from the folder, delete respective database records
    files_to_be_deleted = []
    for i in processed_files_list:
        if i not in files_and_ts:
            files_to_be_deleted.append(i)

    if myexecute:
        for i in files_to_be_deleted:
        
            xlfile = i.split('@@@')[0]
            xltimestamp = i.split('@@@')[1]

            # get existing mrv_id for this file
            stmt = f"select mrv_id from {header_table} where origin_file_name = '{xlfile}'"
            cursor.execute(stmt)
            dbrow = cursor.fetchone()

            mrvid = dbrow[0]

            # delete records for this file
            #stmt = f"delete from {records_table} where mrv_id = {mrvid};"
            #cursor.execute(stmt)
            #stmt = f"delete from {header_table} where mrv_id = {mrvid};"
            #cursor.execute(stmt)

    # process files that don't exist in the processed files list
    files_to_be_processed = []
    for i in files_and_ts:
        if i not in processed_files_list:
            files_to_be_processed.append(i)

    if myexecute:
        for i in files_to_be_processed:

            xlfile = i.split('@@@')[0]
            xltimestamp = i.split('@@@')[1]

            # get existing mrv_id for this file
            stmt = f"select mrv_id from {header_table} where origin_file_name = '{xlfile}'"
            cursor.execute(stmt)
            dbrow = cursor.fetchone()

            try:
                mrvid = dbrow[0]

                # delete records for this file
                #stmt = f"delete from {records_table} where mrv_id = {mrvid};"
                #cursor.execute(stmt)
                #stmt = f"delete from {header_table} where mrv_id = {mrvid};"
                #cursor.execute(stmt)

                # process it
                #process_file(xlfile, xltimestamp, mrvid)

            except:
                #process_file(xlfile, xltimestamp)
                pass
        # end for
    
    # end if myexecute

    return_value = 'Files to be deleted: ' + str(len(files_to_be_deleted)) + ' / Files to be processed: ' + str(len(files_to_be_processed))
    return return_value

    #cnxn.close()

######################################

if st.button('MRV'):
    result = process_mrv()
    st.write('result: %s' % result)
