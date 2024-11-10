# https://pmc.ncbi.nlm.nih.gov/articles/PMC8011774/
# https://icm-experimental.springeropen.com/articles/10.1186/s40635-021-00421-8

from openpyxl import load_workbook
import csv

def pmsa(cvp_rap, map, co, age, height, weight):
  c = 0.038 * (94.17 + 0.193 * age) / (4.5 * (0.99**(age-15)) * 0.007184 * (height**0.725) * (weight**0.425))
  pmsa = 0.96*cvp_rap + 0.04*map + c*co
  pvr = pmsa - cvp_rap
  eh = pvr/pmsa
  return (round(pmsa,2), round(pvr,2), round(eh,2))

data = [['CVP', 'MAP', 'CO', 'AGE', 'HEIGHT', 'WEIGHT', 'PMSA', 'PVR', 'EH']]

wb = load_workbook('thanos.xlsx')
ws = wb.worksheets[0]
for xlrow in ws.iter_rows():
  row = [cell.value for cell in xlrow]
  if row[6] and row[6]!='CVP': # exclude empty rows and header rows
    cvp = float(row[6])
    map = float(row[7])
    co = float(row[8])
    age = float(row[9])
    height = float(row[10])
    weight = float(row[11])
    result = pmsa(cvp, map, co, age, height, weight)
    print(cvp, '\t', map, '\t', co, '\t', age, '\t', height, '\t', weight, '\t', result[0], '\t', result[1], '\t', result[2])
    data.append([cvp, map, co, age, height, weight, result[0], result[1], result[2]])

with open('thanos.csv', 'w', newline='') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerows(data)
